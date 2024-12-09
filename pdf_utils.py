from pdf2image import convert_from_path
import cv2
import numpy as np
import pandas as pd
import pymupdf 
from openpyxl.styles import PatternFill
from datetime import datetime

def utils_group_row_point(values):
    values = sorted(values, key=lambda x: (x[1], x[0]))
    values = [value for value in values if value[1] >= 50]

    groups = [[values[0]]]  

    for i in range(1, len(values)):
        if values[i][1] - values[i-1][1] >= 50:
            groups.append([values[i]])  
        else:
            groups[-1].append(values[i])  

    return groups

def detect_by_color(bgr_color, image):
    lower_bound = bgr_color - 20
    upper_bound = bgr_color + 20

    mask = cv2.inRange(image, lower_bound, upper_bound)

    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    rectangle = []

    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        cv2.rectangle(image, (x, y), (x+w, y+h), (0, 255, 0), 2)
        rectangle.append((x*0.12, y*0.12, (w+x)*0.12, (h+y)*0.12))

    return rectangle

def extract_text_from_bbox(pdf_path, bounding_boxes, bounding_boxes_header):
    doc = pymupdf.open(pdf_path)
    table_dicts = []
    
    row_points = utils_group_row_point(bounding_boxes)
    row_point_headers = utils_group_row_point(bounding_boxes_header)

    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        for n in range(len(row_point_headers)):
            table_dict = {}
            datas = row_points[n]
            headers = row_point_headers[n]
            for header in headers:
                for data in datas:
                    if header[0] - 5 <= data[0] and header[2] + 5 >= data[2]:
                        h = page.get_text("text", clip=pymupdf.Rect(header)).strip()
                        d = page.get_text("text", clip=pymupdf.Rect(data)).strip()
                        if h not in table_dict:
                            table_dict[h] = [d]
                        else:
                            table_dict[h].append(d)
            table_dicts.append(table_dict)

        page_width, page_height = page.rect.width, page.rect.height
        bottom_half_rect = pymupdf.Rect(0, page_height * 0.65, page_width, page_height)
        bottom_text = page.get_text("text", clip=bottom_half_rect)
    return table_dicts, bottom_text

def process_pdf(pdf_path, poppler_path, header_color, data_color):
    """
    Process a PDF to extract text based on specified header and data colors.
    """
    pages = convert_from_path(pdf_path, 600, poppler_path=poppler_path, use_pdftocairo=True)
    page = np.array(pages[0]) 
    
    header_bbox = detect_by_color(header_color, page)
    data_bbox = detect_by_color(data_color, page)
    
    extracted_texts, bottom_text = extract_text_from_bbox(pdf_path, data_bbox, header_bbox)
    
    tables = [pd.DataFrame(extracted_text) for extracted_text in extracted_texts]
    return tables[0], bottom_text
    

def pdf_check(mt_df, df, bottom_text, wb, ws):
  for i, row in df.iterrows():
    row_to_check = mt_df[(mt_df['System Key'] == row['Systemkey'].strip()) | (mt_df['Origin Site ID'] == row['Origin Site ID'].strip())]
    fo_meter_column = [col for col in df.columns if "connection" in col.lower()]
    site_column = [col for col in df.columns if "site" in col.lower()]
    activation_date_column = [col for col in df.columns if "activation" in col.lower()]
    signing_date_column = [col for col in df.columns if "signing" in col.lower()]
    check_dict = {}
    if len(row_to_check) > 0:
        span_id = row['Span ID'].strip().split('-')
        ring_id = "-".join(span_id[:4])
        far_end_id = span_id[5]
        
        span_id_mt_list = [row_to_check['Ring ID'].to_string(index=False), row_to_check['Origin Site ID'].to_string(index=False), row_to_check['Destination Site ID'].to_string(index=False)]
        span_id_mt = '-'.join(span_id_mt_list)
        
        check_dict['System Key'] = row_to_check['System Key'].to_string(index=False) == row['Systemkey'].strip()

        check_dict['Origin Site ID'] = row_to_check['Origin Site ID'].to_string(index=False) == row[site_column[0]].strip()
        
        check_dict['Origin Site Name'] = row_to_check['Origin Site Name'].to_string(index=False) == row[site_column[1]].strip()

        check_dict['Destination Site ID'] = row_to_check['Destination Site ID'].to_string(index=False) == far_end_id

        check_dict['Ring ID'] = row_to_check['Ring ID'].to_string(index=False) == ring_id
 
        date_object_activation = datetime.strptime(row[activation_date_column[0]].strip(), '%d-%b-%y')
        date_object_signing = datetime.strptime(row[signing_date_column[0]].strip(), '%d-%b-%y')
        check_dict['BAA Date (Aktivasi)'] = date_object_activation == row_to_check['BAA Date (Aktivasi)'].iloc[0]
        check_dict['BAA Date (Aktivasi)'] = date_object_signing == row_to_check['BAA Date (Aktivasi)'].iloc[0]

        check_dict['Panjang OTDR (M)'] = float(row_to_check['Panjang OTDR (M)'].to_string(index=False)) == float(row[fo_meter_column[0]].strip())

        check_dict['Span ID'] = span_id_mt == row['Span ID'].strip()
        
        check_dict['Signing Date'] = row[activation_date_column[0]] == row[signing_date_column[0]].strip()

        check_dict['Signing Person'] = "Mochamad Abbari Ramadhona" in bottom_text

        color_cells_based_on_conditions(wb, ws, check_dict, row['Systemkey'].strip())
    else:
        check_dict['BAA'] = False
    return check_dict

def color_cells_based_on_conditions(wb, ws, condition_dict, filter_value):
    """
    Fills Excel cells with red color based on a condition dictionary and filter value.

    Args:
        file_name (str): Path to the input Excel file.
        output_file (str): Path to save the updated Excel file.
        condition_dict (dict): A dictionary where keys are column names and values are booleans.
                               If the value is False, cells in that column will be filled with red.
        filter_value (str): The value to filter rows in the "SYSTEM_KEY" column.
    """

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    column_indices = {key: header_row.index(key) + 1 for key in condition_dict.keys()}

    for row in ws.iter_rows(min_row=2):  
        system_key_value = row[column_indices['System Key'] - 1].value  
        if str(system_key_value) == filter_value:
            for col_name, condition in condition_dict.items():
                if not condition:  
                    cell = row[column_indices[col_name] - 1]
                    cell.fill = red_fill
                elif condition:
                    cell = row[column_indices[col_name] - 1]
                    cell.fill = green_fill
            break

    wb.save("mt_database.xlsx")
